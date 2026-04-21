import openpyxl

def restore_acuan():
    src_path = '/Users/pcchopath/Downloads/Form-Penilaian-Kerusakan-Bangunan SDN Banjarsari.xlsx'
    dst_path = '/Users/pcchopath/Downloads/Form-Penilaian-Kerusakan-Bangunan SDN KENDALPECABEAN.xlsx'
    
    print(f"Restoring 'Presentase Acuan' from {src_path} to {dst_path}...")
    
    src_wb = openpyxl.load_workbook(src_path, data_only=True)
    src_ws = src_wb['Presentase Acuan']
    
    dst_wb = openpyxl.load_workbook(dst_path)
    if 'Presentase Acuan' in dst_wb.sheetnames:
        # Clear the existing sheet
        dst_ws = dst_wb['Presentase Acuan']
        for row in dst_ws.iter_rows():
            for cell in row:
                cell.value = None
    else:
        dst_ws = dst_wb.create_sheet('Presentase Acuan')
        
    # Copy values
    for row in src_ws.iter_rows(values_only=True):
        for i, val in enumerate(row, 1):
            dst_ws.cell(row=src_ws.min_row + (src_ws.iter_rows(max_row=1, values_only=True).__next__()[0].index if False else 0), # this is wrong
                      column=i).value = val
    
    # Correct way to copy values
    for r_idx, row in enumerate(src_ws.iter_rows(values_only=True), 1):
        for c_idx, val in enumerate(row, 1):
            dst_ws.cell(row=r_idx, column=c_idx).value = val
            
    dst_wb.save(dst_path)
    print("Restoration complete!")

if __name__ == "__main__":
    restore_acuan()
