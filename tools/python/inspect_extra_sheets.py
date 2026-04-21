from openpyxl import load_workbook

src_path = '/Users/pcchopath/Downloads/Form-Penilaian-Kerusakan-Bangunan SDN Banjarsari.xlsx'
wb = load_workbook(src_path, data_only=True)

for s_name in ['Atap', 'Template']:
    print(f"--- Sheet: {s_name} ---")
    ws = wb[s_name]
    for row in ws.iter_rows(max_row=50, values_only=True):
        print(row)
    print()
