from openpyxl import load_workbook

target_path = '/Users/pcchopath/Downloads/Form-Penilaian-Kerusakan-Bangunan SDN KENDALPECABEAN.xlsx'
wb = load_workbook(target_path)
ws = wb['R. KELAS 1 & 2']

# Let's check row 19, column W (index 22)
cell_value = ws.cell(row=19, column=23).value # 1-indexed
print(f"Row 19, Col 23 value: {cell_value}")
