import openpyxl
from openpyxl import load_workbook

def set_cell_safe(ws, row, col, value):
    cell = ws.cell(row=row, column=col)
    if isinstance(cell, openpyxl.cell.cell.MergedCell):
        for range_ in ws.merged_cells.ranges:
            if cell.coordinate in range_:
                ws.cell(row=range_.min_row, column=range_.min_col).value = value
                return
    else:
        cell.value = value

def restore_to_sop(file_path, gold_path):
    print(f"Restoring {file_path} to official SOP...")
    
    target_wb = load_workbook(file_path)
    gold_wb = load_workbook(gold_path)
    gold_ws = gold_wb['RUANG KELAS']
    
    skip_sheets = {'Presentase', 'Presentase Acuan', 'Template', 'Sheet1', 'Atap'}
    room_sheets = [s for s in target_wb.sheetnames if s not in skip_sheets]
    
    for sheet_name in room_sheets:
        print(f"Restoring SOP for {sheet_name}...")
        ws = target_wb[sheet_name]
        
        # 1. Restore formulas for P:Y based on gold template's exact strings
        for row in range(19, 41):
            # Restore P:V (Columns 16-22)
            for col in range(16, 23):
                col_letter = openpyxl.utils.get_column_letter(col)
                i_col_letter = openpyxl.utils.get_column_letter(col - 7)
                weight_col_letter = openpyxl.utils.get_column_letter(col)
                # EXACT FORMULA FROM GOLD
                set_cell_safe(ws, row, col, f"={i_col_letter}{row}/G{row}*${weight_col_letter}$17")
            
            # Restore W (23)
            set_cell_safe(ws, row, 23, f"=SUM(P{row}:V{row})")
            
            # Restore Y (25)
            set_cell_safe(ws, row, 25, f"=W{row}*X{row}")
            
        # 2. Fix Page Setup for PDF (as per gold template)
        ws.print_area = f"'{sheet_name}'!$B$3:$Y$76"
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0 # Auto height
        
    target_wb.save(file_path)
    print("SOP Restoration complete!")

if __name__ == "__main__":
    target = '/Users/pcchopath/Downloads/Form-Penilaian-Kerusakan-Bangunan SDN KENDALPECABEAN.xlsx'
    gold = '/Users/pcchopath/Downloads/Form-Penilaian-Kerusakan-Bangunan SDN Banjarsari.xlsx'
    restore_to_sop(target, gold)
