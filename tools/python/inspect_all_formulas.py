from openpyxl import load_workbook

src_path = '/Users/pcchopath/Downloads/Form-Penilaian-Kerusakan-Bangunan SDN Banjarsari.xlsx'
wb = load_workbook(src_path, data_only=False)
ws = wb['RUANG KELAS']

for col in range(16, 26): # P to Z
    val = ws.cell(row=20, column=col).value
    print(f"Col {col}: {val}")
