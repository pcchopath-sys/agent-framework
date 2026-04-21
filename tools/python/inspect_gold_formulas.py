from openpyxl import load_workbook

src_path = '/Users/pcchopath/Downloads/Form-Penilaian-Kerusakan-Bangunan SDN Banjarsari.xlsx'
wb = load_workbook(src_path, data_only=False)
ws = wb['RUANG KELAS']

print("Inspecting I20:O31 in gold template for FORMULAS/RAW VALUES...")
for row in range(20, 32):
    row_data = []
    for col in range(9, 16): # Col I to O
        val = ws.cell(row=row, column=col).value
        row_data.append(val)
    print(f"Row {row}: {row_data}")
