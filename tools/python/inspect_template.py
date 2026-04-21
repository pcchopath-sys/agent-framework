from openpyxl import load_workbook

target_path = '/Users/pcchopath/Downloads/Form-Penilaian-Kerusakan-Bangunan SDN KENDALPECABEAN.xlsx'
wb = load_workbook(target_path)
ws = wb['Template']

for row in ws.iter_rows(max_row=20, values_only=True):
    print(row)
