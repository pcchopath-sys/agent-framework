from openpyxl import load_workbook

gold_path = '/Users/pcchopath/Downloads/Form-Penilaian-Kerusakan-Bangunan SDN Banjarsari.xlsx'
wb = load_workbook(gold_path)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"Sheet: {sheet_name}")
    print(f"  Validations object: {ws.data_validations}")
    for dv in ws.data_validations.dataValidation:
        print(f"  Formula: {dv.formula1}, Range: {dv.sqref}")
