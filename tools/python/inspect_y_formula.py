from openpyxl import load_workbook

target_path = '/Users/pcchopath/Downloads/Form-Penilaian-Kerusakan-Bangunan SDN KENDALPECABEAN.xlsx'
wb = load_workbook(target_path, data_only=False)
ws = wb['R. KELAS 1 & 2']

print("Inspecting formula in Y20...")
print(ws.cell(row=20, column=25).value)
