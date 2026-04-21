import openpyxl
from openpyxl.styles import Alignment, Border, Side, Font
from openpyxl.worksheet.page import PageMargins

def set_cell_safe(ws, row, col, value):
    cell = ws.cell(row=row, column=col)
    if isinstance(cell, openpyxl.cell.cell.MergedCell):
        for range_ in ws.merged_cells.ranges:
            if cell.coordinate in range_:
                ws.cell(row=range_.min_row, column=range_.min_col).value = value
                return
    else:
        cell.value = value

def mirror_layout(src_ws, dst_ws):
    # Mirror Column Widths
    for col, dim in src_ws.column_dimensions.items():
        dst_ws.column_dimensions[col].width = dim.width
    
    # Mirror Row Heights
    for row, dim in src_ws.row_dimensions.items():
        dst_ws.row_dimensions[row].height = dim.height
    
    # Clear Merged Cells in dst_ws first
    merged_ranges = list(dst_ws.merged_cells.ranges)
    for r in merged_ranges:
        dst_ws.unmerge_cells(str(r))
    
    # Mirror Merged Cells from src_ws
    for merged_range in src_ws.merged_cells.ranges:
        dst_ws.merge_cells(str(merged_range))

def apply_damage_logic(ws, sheet_name):
    print(f"Processing {sheet_name}...")
    
    damage_map = {
        22: {"H": "Rangka atap rubuh/lapuk berat dimakan rayap", "level": 5, "vol": 1},
        24: {"H": "Plafon rubuh/berlubang besar di banyak titik", "level": 4, "vol": 1},
        29: {"H": "Plafond mengalami kerusakan berat akibat rembesan air hujan", "level": 4, "vol": 1},
    }
    
    others = [20, 21, 23, 25, 26, 27, 28, 30, 31, 34]
    no_damage = [19, 32, 33]
    
    for row_idx in range(19, 41):
        # Clear I:O range (Col 9 to 15) to ensure no text
        for col_idx in range(9, 16):
            set_cell_safe(ws, row_idx, col_idx, None)

        if row_idx in damage_map:
            data = damage_map[row_idx]
            set_cell_safe(ws, row_idx, 8, data["H"])
            # Column I (9) as indicator: 1 for damaged
            set_cell_safe(ws, row_idx, 9, 1)
            # Map level 1-5 to Col J-N (10-14)
            col_idx = 9 + data["level"]
            if col_idx <= 15:
                set_cell_safe(ws, row_idx, col_idx, data["vol"])
            set_cell_safe(ws, row_idx, 24, 0.8 if data["level"] >= 4 else 0.4)
        elif row_idx in others:
            set_cell_safe(ws, row_idx, 8, "Ada kerusakan, namun diindikasi tidak membahayakan keselamatan pemanfaatan ruang/bangunan")
            set_cell_safe(ws, row_idx, 9, 1)
            set_cell_safe(ws, row_idx, 12, 1) # Ringan (Level 2 -> Col L)
            set_cell_safe(ws, row_idx, 24, 0.05)
        elif row_idx in no_damage:
            set_cell_safe(ws, row_idx, 8, "Tidak ada kerusakan")
            set_cell_safe(ws, row_idx, 9, 0)
            set_cell_safe(ws, row_idx, 24, 0.01)
        else:
            set_cell_safe(ws, row_idx, 8, "Diindikasi dalam kondisi baik")
            set_cell_safe(ws, row_idx, 9, 0)
            set_cell_safe(ws, row_idx, 24, 0.01)


def setup_print(ws):
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_area = 'A1:Z40'

def main():
    src_path = '/Users/pcchopath/Downloads/Form-Penilaian-Kerusakan-Bangunan SDN Banjarsari.xlsx'
    dst_path = '/Users/pcchopath/Downloads/Form-Penilaian-Kerusakan-Bangunan SDN KENDALPECABEAN.xlsx'
    
    src_wb = openpyxl.load_workbook(src_path)
    src_ws = src_wb['RUANG KELAS']
    
    dst_wb = openpyxl.load_workbook(dst_path)
    target_sheets = ['R. KELAS 1 & 2', 'R. KELAS 3 & 5B', 'KAMAR MANDI SISWA', 'R. KELAS 4', 'R. KM SISWI', 'RUANG GURU ']

    
    for sheet_name in target_sheets:
        if sheet_name not in dst_wb.sheetnames:
            print(f"Sheet {sheet_name} not found, skipping.")
            continue
            
        ws = dst_wb[sheet_name]
        mirror_layout(src_ws, ws)
        apply_damage_logic(ws, sheet_name)
        setup_print(ws)
        
    dst_wb.save(dst_path)
    print("Done!")

if __name__ == "__main__":
    main()
