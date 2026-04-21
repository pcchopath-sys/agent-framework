from openpyxl import load_workbook

src_path = '/Users/pcchopath/Downloads/Form-Penilaian-Kerusakan-Bangunan SDN Banjarsari.xlsx'
wb = load_workbook(src_path, data_only=True)
ws = wb['RUANG KELAS']

print("Studying Damage Criteria in Master Template...")
# Inspect rows 35-45 where conclusion and thresholds are usually located
for row in range(35, 46):
    row_data = []
    for col in range(1, 31):
        val = ws.cell(row=row, column=col).value
        if val:
            row_data.append((col, val))
    if row_data:
        print(f"Row {row}: {row_data}")
