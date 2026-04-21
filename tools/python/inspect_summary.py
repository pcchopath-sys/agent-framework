from openpyxl import load_workbook

target_path = '/Users/pcchopath/Downloads/Form-Penilaian-Kerusakan-Bangunan SDN KENDALPECABEAN.xlsx'
wb = load_workbook(target_path, data_only=False)
ws = wb['R. KELAS 1 & 2']

print(f"H35: {ws.cell(row=35, column=8).value}")
print(f"Y35: {ws.cell(row=35, column=25).value}")
