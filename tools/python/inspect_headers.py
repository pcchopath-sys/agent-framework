from openpyxl import load_workbook

src_path = '/Users/pcchopath/Downloads/Form-Penilaian-Kerusakan-Bangunan SDN Banjarsari.xlsx'
wb = load_workbook(src_path, data_only=True)
ws = wb['RUANG KELAS']

for col in range(9, 16): # I to O
    print(f"Col {col}: {ws.cell(row=15, column=col).value}")
