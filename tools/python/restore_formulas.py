import openpyxl
from openpyxl import load_workbook

def set_cell_safe(ws, row, col, value):
    cell = ws.cell(row=row, column=col)
    if isinstance(cell, openpyxl.cell.cell.MergedCell):
        for range_ in ws.merged_cells.ranges:
            if cell.coordinate in range_:
                ws.cell(row=range_.min_row, column=range_.min_col).value = value
                return
    else:
        cell.value = value

def restore_formulas(file_path):
    print(f"Restoring formulas in {file_path}...")
    wb = load_workbook(file_path)
    
    skip_sheets = {'Presentase', 'Presentase Acuan', 'Template', 'Sheet1', 'Atap'}
    room_sheets = [s for s in wb.sheetnames if s not in skip_sheets]
    
    for sheet_name in room_sheets:
        print(f"Processing {sheet_name}...")
        ws = wb[sheet_name]
        
        for row in range(19, 41):
            # Ensure G (Total Volume) is not empty to avoid #DIV/0!
            g_val = ws.cell(row=row, column=7).value
            if g_val is None or g_val == 0:
                set_cell_safe(ws, row, 7, 10) # Default total volume 10
                
            # Restore P:V
            for col in range(16, 23):
                col_letter = openpyxl.utils.get_column_letter(col)
                i_col_letter = openpyxl.utils.get_column_letter(col - 7)
                weight_col_letter = openpyxl.utils.get_column_letter(col)
                set_cell_safe(ws, row, col, f"={i_col_letter}{row}/G{row}*${weight_col_letter}$17")
            
            # Restore W (23)
            set_cell_safe(ws, row, 23, f"=SUM(P{row}:V{row})")
            
            # Restore Y (25)
            set_cell_safe(ws, row, 25, f"=W{row}*X{row}")
            
    wb.save(file_path)
    print("Formulas restored successfully!")

if __name__ == "__main__":
    target_file = '/Users/pcchopath/Downloads/Form-Penilaian-Kerusakan-Bangunan SDN KENDALPECABEAN.xlsx'
    restore_formulas(target_file)
