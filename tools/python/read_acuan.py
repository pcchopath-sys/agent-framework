from openpyxl import load_workbook

gold_path = '/Users/pcchopath/Downloads/Form-Penilaian-Kerusakan-Bangunan SDN Banjarsari.xlsx'
wb = load_workbook(gold_path, data_only=True)
ws = wb['Presentase Acuan']

print("Contents of Presentase Acuan:")
for row in ws.iter_rows(values_only=True):
    print(row)
