from openpyxl import load_workbook

target_path = '/Users/pcchopath/Downloads/Form-Penilaian-Kerusakan-Bangunan SDN KENDALPECABEAN.xlsx'
wb = load_workbook(target_path, read_only=True)
print(f"Sheets in target: {wb.sheetnames}")
