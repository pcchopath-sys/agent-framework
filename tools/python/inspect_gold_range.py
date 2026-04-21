from openpyxl import load_workbook

src_path = '/Users/pcchopath/Downloads/Form-Penilaian-Kerusakan-Bangunan SDN Banjarsari.xlsx'
wb = load_workbook(src_path, data_only=True)
ws = wb['RUANG KELAS']

print("Inspecting I20:O31 in gold template...")
for row in range(20, 32):
    row_data = []
    for col in range(9, 16): # Col I to O
        row_data.append(ws.cell(row=row, column=col).value)
    print(f"Row {row}: {row_data}")
