from openpyxl import load_workbook

target_path = '/Users/pcchopath/Downloads/Form-Penilaian-Kerusakan-Bangunan SDN KENDALPECABEAN.xlsx'
wb = load_workbook(target_path, data_only=True)
ws = wb['R. KELAS 1 & 2']

for row in ws.iter_rows(max_row=20, values_only=True):
    print(row)
