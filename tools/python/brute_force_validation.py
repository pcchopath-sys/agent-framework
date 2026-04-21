import re
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation

def apply_brute_force_validation(file_path):
    print(f"Opening {file_path}...")
    wb = load_workbook(file_path)
    
    # Define sheets to skip
    skip_sheets = {'Presentase', 'Presentase Acuan', 'Template', 'Sheet1', 'Atap'}
    room_sheets = [s for s in wb.sheetnames if s not in skip_sheets]
    
    for sheet_name in room_sheets:
        print(f"Processing sheet: {sheet_name}")
        ws = wb[sheet_name]
        
        # We'll track validations to avoid adding the same one multiple times
        # openpyxl's add_data_validation adds to a list; we should be careful.
        # Actually, creating a new DV object per range is safer.
        
        for row in range(19, 101):
            # Column 23 (W) contains the VLOOKUP formula
            cell_w = ws.cell(row=row, column=23)
            formula = cell_w.value
            
            if formula and isinstance(formula, str) and 'VLOOKUP' in formula:
                # Extract range: look for 'Presentase Acuan'!Bxx:Cxx
                match = re.search(r"('Presentase Acuan'!B\d+:C\d+)", formula)
                if match:
                    full_range = match.group(1)
                    # Convert Bxx:Cxx to Bxx:Bxx for the dropdown
                    # Example: 'Presentase Acuan'!B12:C18 -> 'Presentase Acuan'!B12:B18
                    dropdown_range = full_range.replace(':C', ':B')
                    
                    # Create DataValidation object
                    dv = DataValidation(type="list", formula1=f"={dropdown_range}", allow_blank=True)
                    ws.add_data_validation(dv)
                    
                    # Apply to Column 9 (I)
                    cell_i = ws.cell(row=row, column=9)
                    dv.add(cell_i.coordinate)
                    
    wb.save(file_path)
    print("Done!")

if __name__ == "__main__":
    target_file = '/Users/pcchopath/Downloads/Form-Penilaian-Kerusakan-Bangunan SDN KENDALPECABEAN.xlsx'
    apply_brute_force_validation(target_file)
