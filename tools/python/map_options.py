import openpyxl
from openpyxl import load_workbook
import json

def map_options():
    target_path = '/Users/pcchopath/Downloads/Form-Penilaian-Kerusakan-Bangunan SDN KENDALPECABEAN.xlsx'
    wb = load_workbook(target_path, data_only=True)
    
    if 'Presentase Acuan' not in wb.sheetnames:
        print("CRITICAL: 'Presentase Acuan' sheet missing!")
        return

    acuan_ws = wb['Presentase Acuan']
    valid_options = set()
    for row in acuan_ws.values:
        for cell in row:
            if cell and isinstance(cell, str):
                valid_options.add(cell)

    skip_sheets = {'Presentase', 'Presentase Acuan', 'Template', 'Sheet1', 'Atap'}
    room_sheets = [s for s in wb.sheetnames if s not in skip_sheets]

    invalid_options = set()
    for sheet_name in room_sheets:
        ws = wb[sheet_name]
        for row in range(19, 41):
            val_i = ws.cell(row=row, column=9).value
            if val_i and val_i not in valid_options:
                invalid_options.add(val_i)

    mapping_data = {
        "invalid_found": list(invalid_options),
        "valid_options": list(valid_options)
    }
    
    with open('options_mapping.json', 'w') as f:
        json.dump(mapping_data, f, indent=2)
    
    print(f"Found {len(invalid_options)} unique invalid options. Saved to options_mapping.json")

if __name__ == "__main__":
    map_options()
