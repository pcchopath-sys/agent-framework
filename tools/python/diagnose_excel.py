import openpyxl
from openpyxl import load_workbook

def diagnose():
    target_path = '/Users/pcchopath/Downloads/Form-Penilaian-Kerusakan-Bangunan SDN KENDALPECABEAN.xlsx'
    acuan_path = target_path # acuan is a sheet in the same file
    
    wb = load_workbook(target_path, data_only=True)
    
    if 'Presentase Acuan' not in wb.sheetnames:
        print("CRITICAL: 'Presentase Acuan' sheet missing!")
        return

    acuan_ws = wb['Presentase Acuan']
    # Extract valid options from Column B of Presentase Acuan
    valid_options = set()
    for row in acuan_ws.iter_rows(min_row=1, max_col=2, values_only=True):
        if row[1]: # Col B
            valid_options.add(row[1])

    skip_sheets = {'Presentase', 'Presentase Acuan', 'Template', 'Sheet1', 'Atap'}
    room_sheets = [s for s in wb.sheetnames if s not in skip_sheets]

    print(f"Auditing {len(room_sheets)} room sheets...\n")
    
    total_errors = 0
    for sheet_name in room_sheets:
        ws = wb[sheet_name]
        print(f"--- Sheet: {sheet_name} ---")
        sheet_has_error = False
        
        # Check rows 19 to 40 (typical damage assessment range)
        for row in range(19, 41):
            val_h = ws.cell(row=row, column=8).value # Col H
            val_i = ws.cell(row=row, column=9).value # Col I
            val_x = ws.cell(row=row, column=24).value # Col X
            
            # 1. Check for empty critical cells
            if not val_h or val_x is None:
                print(f"Row {row}: Missing critical data (H={val_h}, X={val_x})")
                sheet_has_error = True
                total_errors += 1
            elif val_i is None and val_h not in ["Tidak ada kerusakan", "Diindikasi dalam kondisi baik"]:
                print(f"Row {row}: Missing damage value in Col I despite damage in H")
                sheet_has_error = True
                total_errors += 1

                
            # 2. Check if value in Col I is in the valid options list
            if val_i and val_i not in valid_options:
                print(f"Row {row}: Invalid option in Col I: '{val_i}'")
                sheet_has_error = True
                total_errors += 1
                
        if not sheet_has_error:
            print("OK")
        print()

    print(f"\nTotal errors found: {total_errors}")

if __name__ == "__main__":
    diagnose()
