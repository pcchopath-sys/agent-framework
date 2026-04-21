from openpyxl import load_workbook

target_path = '/Users/pcchopath/Downloads/Form-Penilaian-Kerusakan-Bangunan SDN KENDALPECABEAN.xlsx'
wb = load_workbook(target_path, data_only=True)
ws = wb['Presentase Acuan']

for row in ws.iter_rows(max_row=100, values_only=True):
    print(row)
