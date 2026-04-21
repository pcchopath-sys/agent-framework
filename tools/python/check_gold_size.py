from openpyxl import load_workbook

src_path = '/Users/pcchopath/Downloads/Form-Penilaian-Kerusakan-Bangunan SDN Banjarsari.xlsx'
wb = load_workbook(src_path, data_only=True)
ws = wb['Presentase Acuan']
print(f"Max row: {ws.max_row}, Max col: {ws.max_column}")
