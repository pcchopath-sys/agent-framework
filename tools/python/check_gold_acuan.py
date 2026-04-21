from openpyxl import load_workbook

src_path = '/Users/pcchopath/Downloads/Form-Penilaian-Kerusakan-Bangunan SDN Banjarsari.xlsx'
wb = load_workbook(src_path, data_only=True)
ws = wb['Presentase Acuan']

for row in ws.iter_rows(max_row=200, values_only=True):
    print(row)
